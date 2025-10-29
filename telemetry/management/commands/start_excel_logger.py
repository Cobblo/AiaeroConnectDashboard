# telemetry/management/commands/start_excel_logger.py
import os
import re
import time
from datetime import datetime
from pathlib import Path

from django.core.management.base import BaseCommand
from django.utils import timezone

from openpyxl import Workbook, load_workbook

from telemetry.models import CurrentVital, Person

# How often to write rows (seconds). Change to 60, 300, etc. as you like.
POLL_SECONDS = 60

def safe_name(s: str) -> str:
    """Make a safe filename from a person name."""
    s = re.sub(r"[^\w\-]+", "_", s.strip())
    return s or "unknown"

def ensure_sheet_with_headers(xlsx_path: Path):
    """
    Create a workbook with a single 'data' sheet with headers if file does not exist.
    Return (workbook, worksheet).
    """
    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
        ws = wb.active
        return wb, ws

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append([
        "date", "time", "person", "device",
        "heart_rate", "spo2", "temp_c",
        "bp_sys", "bp_dia",
        "battery_pct", "rssi",
        "ts_iso"
    ])
    wb.save(xlsx_path)
    return wb, wb.active

def row_from_cv(p: Person, cv: CurrentVital):
    local_ts = timezone.localtime(cv.ts)
    return [
        local_ts.strftime("%Y-%m-%d"),
        local_ts.strftime("%H:%M:%S"),
        p.name,
        cv.device.device_id if cv.device else "",
        cv.heart_rate or "",
        cv.spo2 or "",
        cv.temp_c or "",
        cv.bp_sys or "",
        cv.bp_dia or "",
        cv.battery_pct or "",
        cv.rssi or "",
        local_ts.isoformat(),
    ]

class Command(BaseCommand):
    help = "Continuously logs each soldier's latest vitals to per-person Excel files, grouped by date."

    def add_arguments(self, parser):
        parser.add_argument(
            "--interval",
            type=int,
            default=POLL_SECONDS,
            help="Polling interval in seconds (default: 60)",
        )
        parser.add_argument(
            "--out",
            type=str,
            default="exports/vitals",
            help="Base output folder (default: exports/vitals)",
        )

    def handle(self, *args, **opts):
        base_out = Path(opts["out"])
        interval = int(opts["interval"])

        # Keep the latest timestamp written for each device to avoid duplicates
        last_written_ts = {}

        self.stdout.write(self.style.SUCCESS(
            f"Excel logger started. Interval={interval}s, output={base_out}"
        ))
        self.stdout.write("Press Ctrl+C to stop.\n")

        while True:
            try:
                # Pull all persons that have a device and a current vital row
                persons = (
                    Person.objects.select_related("device")
                    .filter(device__isnull=False)
                    .order_by("id")
                )

                for p in persons:
                    dev = p.device
                    # There might not yet be a CurrentVital row
                    try:
                        cv = dev.current
                    except CurrentVital.DoesNotExist:
                        continue
                    if not cv or not cv.ts:
                        continue

                    # Skip if we already wrote this timestamp for the device
                    last_ts = last_written_ts.get(dev.id)
                    if last_ts is not None and cv.ts <= last_ts:
                        continue

                    # Compute output path: exports/vitals/YYYY-MM-DD/Person_Name.xlsx
                    day = timezone.localtime(cv.ts).strftime("%Y-%m-%d")
                    out_dir = base_out / day
                    xlsx_path = out_dir / f"{safe_name(p.name)}.xlsx"

                    # Ensure workbook & headers, then append one row
                    wb, ws = ensure_sheet_with_headers(xlsx_path)
                    ws.append(row_from_cv(p, cv))
                    wb.save(xlsx_path)

                    last_written_ts[dev.id] = cv.ts
                    self.stdout.write(
                        f"✔ wrote {p.name} @ {timezone.localtime(cv.ts).strftime('%H:%M:%S')} -> {xlsx_path}"
                    )

            except Exception as e:
                # Don't crash the loop if one write fails.
                self.stderr.write(self.style.ERROR(f"Error: {e}"))

            time.sleep(interval)
