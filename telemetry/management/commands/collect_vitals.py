import json
from pathlib import Path
from datetime import datetime
from django.core.management.base import BaseCommand
from django.conf import settings
from django.utils import timezone
import requests

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

from telemetry.models import Soldier, VitalReading


def append_row_xlsx(xlsx_path: Path, sheet_name: str, header: list, row: list):
    if xlsx_path.exists():
        wb = load_workbook(xlsx_path)
    else:
        wb = Workbook()
        # openpyxl creates a default sheet; we'll remove it after first save if needed

    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(title=sheet_name)

    if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
        ws.append(header)
        # make header bold & auto-col width-ish
        for i, h in enumerate(header, start=1):
            ws.cell(row=1, column=i).font = ws.cell(row=1, column=i).font.copy(bold=True)
            ws.column_dimensions[get_column_letter(i)].width = max(12, len(str(h)) + 2)

    ws.append(row)

    # remove the default "Sheet" if it’s empty and not our target
    if "Sheet" in wb.sheetnames and wb["Sheet"].max_row == 1 and wb["Sheet"]["A1"].value is None:
        std = wb["Sheet"]; wb.remove(std)

    wb.save(xlsx_path)


class Command(BaseCommand):
    help = "Fetch current vitals and append them to DB and today's Excel workbook."

    def handle(self, *args, **kwargs):
        # 1) FETCH CURRENT from your API (replace URL when real API is ready)
        url = "http://127.0.0.1:8000/api/current/"  # change to your gateway if needed
        try:
            r = requests.get(url, timeout=10)
            r.raise_for_status()
            data = r.json()
            items = data.get("items", [])
        except Exception as e:
            self.stderr.write(self.style.ERROR(f"Failed to fetch current: {e}"))
            return

        now = timezone.now()

        # 2) Ensure Soldiers exist & write readings to DB
        created_count = 0
        for it in items:
            sid = int(it.get("person_id") or 0) or None
            name = it.get("person") or f"Soldier {sid or '?'}"
            device_id = it.get("device_id") or f"NODE_{sid or 'XX'}"

            if sid is None:
                # skip items without a person_id (we need it as PK)
                continue

            soldier, _ = Soldier.objects.get_or_create(
                id=sid,
                defaults={"name": name, "device_id": device_id},
            )
            # keep name/device fresh if changed
            if soldier.name != name or soldier.device_id != device_id:
                soldier.name = name
                soldier.device_id = device_id
                soldier.save(update_fields=["name", "device_id"])

            vr = VitalReading.objects.create(
                soldier=soldier, ts=now,
                hr=it.get("hr"), spo2=it.get("spo2"), temp=it.get("temp"),
                bp_sys=it.get("bp_sys"), bp_dia=it.get("bp_dia"),
                battery=it.get("battery"), rssi=it.get("rssi"),
            )
            created_count += 1

        # 3) Append to Excel workbook (one workbook per day, one sheet per soldier)
        book_name = f"vitals_{now.strftime('%Y-%m-%d')}.xlsx"
        xlsx_path = Path(getattr(settings, "EXCEL_DIR")).joinpath(book_name)

        header = ["Timestamp", "Soldier ID", "Name", "Device ID", "HR", "SpO2", "Temp", "BP_SYS", "BP_DIA", "Battery", "RSSI"]
        # get latest readings we just inserted (by ts == now)
        rows = (
            VitalReading.objects
            .filter(ts__gte=now.replace(second=0, microsecond=0))
            .select_related("soldier")
            .order_by("soldier_id")
        )

        for r in rows:
            sheet = f"{r.soldier.name}".strip()[:31]  # Excel sheet name limit
            append_row_xlsx(
                xlsx_path,
                sheet,
                header,
                [
                    r.ts.astimezone(timezone.get_current_timezone()).strftime("%Y-%m-%d %H:%M:%S"),
                    r.soldier_id, r.soldier.name, r.soldier.device_id,
                    r.hr, r.spo2, r.temp, r.bp_sys, r.bp_dia, r.battery, r.rssi
                ]
            )

        self.stdout.write(self.style.SUCCESS(
            f"Stored {created_count} reading(s), appended to {xlsx_path.name}"
        ))
